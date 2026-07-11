import os
import pandas as pd
import qrcode

from PIL import Image
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


def generate_qr(
    excel_file,
    logo_path,
    output_folder,
    output_filename,
    url_template,
    status_callback=None,
    progress_callback=None,
    count_callback=None
):
    """ 
    Membuat QR APAR dan file Excel hasil..
    """
    if not excel_file:
        raise ValueError("File Excel belum dipilih.")

    if not logo_path:
        raise ValueError("Logo belum dipilih.")

    if not output_folder:
        raise ValueError("Folder output belum dipilih.")

    if not output_filename:
        raise ValueError("Nama file output belum diisi.")
    
    if not url_template:
        raise ValueError("Base URL belum diisi.")

    if not os.path.isfile(excel_file):
        raise FileNotFoundError(
            f"File Excel tidak ditemukan:\n{excel_file}"
        )

    if not os.path.isfile(logo_path):
        raise FileNotFoundError(
            f"Logo tidak ditemukan:\n{logo_path}"
        )

    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    qr_folder = os.path.join(output_folder, "QR Images")

    os.makedirs(qr_folder, exist_ok=True)

    output_file = os.path.join(
        output_folder,
        output_filename
    )

    ##################################################

    df = pd.read_excel(excel_file)

    logo = Image.open(
        logo_path
    ).convert("RGBA")

    hyperlinks = []

    qr_files = []

    total = len(df)

    ##################################################

    for index, row in df.iterrows():

        link = url_template

        for kolom in df.columns:
            nilai = str(row[kolom]).strip()

            link = link.replace(
                f"{{{kolom}}}",
                nilai
            )
        if status_callback:
            status_callback(
                f"Membuat QR {index+1}/{total}"
            )
        if count_callback:
            count_callback(
                f"{index+1} / {total} QR"
            )

        hyperlinks.append(link)

        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=20,
            border=1
        )

        qr.add_data(link)

        qr.make(fit=True)

        qr_image = qr.make_image(
            fill_color="black",
            back_color="white"
        ).convert("RGBA")

        ##################################################

        logo_resize = logo.resize(
            (300, 300),
            Image.LANCZOS
        )

        padding = 15

        background = Image.new(
            "RGBA",
            (
                logo_resize.width + padding * 2,
                logo_resize.height + padding * 2
            ),
            (255, 255, 255, 255)
        )

        background.paste(
            logo_resize,
            (padding, padding),
            logo_resize
        )

        position = (
            (qr_image.width - background.width) // 2,
            (qr_image.height - background.height) // 2
        )

        qr_image.paste(
            background,
            position
        )

        ##################################################

        kode_file = str(row[df.columns[1]]).strip()

        for c in r'\/:*?"<>|':
            kode_file = kode_file.replace(c, "_")

        nama_file = f"{kode_file}.png"

        lokasi_file = os.path.join(
            qr_folder,
            nama_file
        )

        qr_image.save(lokasi_file)

        qr_files.append(lokasi_file)

        ##################################################

        if progress_callback:
            progress = int(
                ((index + 1) / total) * 100
            )

            progress_callback(progress)

        print(f"Berhasil membuat QR : {kode_file}")

    ##################################################

    df["Hyperlink"] = hyperlinks

    df["QR File"] = qr_files

    df.to_excel(
        output_file,
        index=False
    )

    ##################################################

    wb = load_workbook(output_file)

    ws = wb.active

    ws["E1"] = "Hasil QR"

    for i, path in enumerate(qr_files, start=2):

        img = XLImage(path)

        img.width = 120
        img.height = 120

        ws.add_image(
            img,
            f"E{i}"
        )

        ws.row_dimensions[i].height = 95

    ws.column_dimensions["E"].width = 22

    wb.save(output_file)

    ##################################################

    if status_callback:
        status_callback(
            f"Selesai. {total} QR berhasil dibuat."
        )

    return output_file

if __name__ == "__main__":

    hasil = generate_qr(
        excel_file="kode_apar.xlsx",
        logo_path="logo_ehs.png",
        output_folder="hasil",
        output_filename="kode_apar_hasil.xlsx",
        url_template="https://contoh.com?id={Kode}"
    )

    print("--------------------------------")
    print("Selesai")
    print("Output :", hasil)